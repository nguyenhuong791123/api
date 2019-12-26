# -*- coding: UTF-8 -*-
import os
import json
from copy import deepcopy
from flask import Blueprint, request, make_response, render_template
import pyexcel as pe
import openpyxl
from openpyxl.chart import BarChart, BarChart3D, LineChart, LineChart3D, PieChart, PieChart3D, AreaChart, AreaChart3D, Reference
from openpyxl.chart.series import DataPoint

from utils.cm.agent import parse_http_accept_language
from utils.cm.files import get_dir, delete_dir, make_dir_local
from utils.viewer.excel import sum_row, sum_cols, sum_array
from utils.viewer.func import EXCEL_FUNC, is_variable_exist

app = Blueprint('excelapi', __name__)

@app.route('/viewer', methods=[ 'GET' ])
def excel_viewer():
    path = 'data/viewer/report_00.xlsx'
    dir = get_dir(None)
    html = 'excel/' + dir + '/viewer.handsontable.html'
    tdir = 'templates/excel/' + dir
    tdir = make_dir_local(tdir)

    try:
        book = openpyxl.load_workbook(path)
        sheetnames = book.get_sheet_names()
        sIdx = 0
        for sn in book.sheetnames:
            sheet = book[sn]
            rIdx = 0
            maxr = len(list(sheet.rows))
            maxc = 0
            sa = []
            for row in sheet.rows:
                if rIdx >= (maxr - 1):
                    for r in range(len(row)):
                        if r > 0 and is_variable_exist(EXCEL_FUNC, row[r].value):
                            row[r].value = sum_cols(sheet, r)

                    categories = Reference(sheet, min_col=1, max_col=1, min_row=maxr, max_row=maxr)
                    datas = Reference(sheet, min_col=2, max_col=(maxc - 1), min_row=2, max_row=(maxr - 1))
                    chart = BarChart()
                    chart.height = 10
                    chart.title = sheetnames[sIdx]
                    chart.add_data(datas, titles_from_data = True)
                    chart.set_categories(categories)
                    sheet.add_chart(chart, "B" + str((maxr + 2)))

                    # chart3d = BarChart3D()
                    # chart3d.legend = None
                    # chart3d.title = sheetnames[sIdx]
                    # chart3d.add_data(datas, titles_from_data = True)
                    # chart3d.set_categories(categories)
                    # sheet.add_chart(chart3d, "K" + str((maxr + 2)))

                    categories =  Reference(sheet, min_col=1, min_row=2, max_row=maxr)
                    datas =  Reference(sheet, min_col=2, max_col=(maxc - 1), min_row=1, max_row=(maxr - 1))
                    line = LineChart()
                    line.height = 10
                    line.title = sheetnames[sIdx]
                    line.add_data(datas, titles_from_data = True)
                    line.set_categories(categories)
                    for ser in range(len(line.series)):
                        s = line.series[ser]
                        s.smooth = True
                        # s.marker.symbol = "triangle"
                        s.marker.graphicalProperties.solidFill = "FF0000"
                        s.marker.graphicalProperties.line.solidFill = "FF0000"
                    sheet.add_chart(line, "B" + str((maxr*2 + (15))))

                    line3d = LineChart3D()
                    line3d.height = 10
                    line3d.legend = None
                    line3d.title = sheetnames[sIdx]
                    line3d.add_data(datas, titles_from_data = True)
                    line3d.set_categories(categories)
                    sheet.add_chart(line3d, "K" + str((maxr*2 + (15))))

                    categories = Reference(sheet, min_col=1, min_row=2, max_row=maxr)
                    datas = Reference(sheet, min_col=2, max_col=(maxc - 1), min_row=1, max_row=(maxr - 1))
                    pie = PieChart()
                    pie.height = 10
                    pie.title = sheetnames[sIdx]
                    pie.add_data(datas, titles_from_data = True)
                    pie.set_categories(categories)
                    slice = DataPoint(idx=0, explosion=20)
                    pie.series[0].data_points = [slice]
                    sheet.add_chart(pie, "B" + str((maxr*3 + (15*2))))

                    pie3d = PieChart3D()
                    pie3d.height = 10
                    pie3d.title = sheetnames[sIdx]
                    pie3d.add_data(datas, titles_from_data = True)
                    pie3d.set_categories(categories)
                    sheet.add_chart(pie3d, "K" + str((maxr*3 + (15*2))))

                    categories = Reference(sheet, min_col=1, min_row=2, max_row=maxr)
                    datas = Reference(sheet, min_col=2, min_row=1, max_col=(maxc - 1), max_row=(maxr - 1))
                    area = AreaChart()
                    area.height = 10
                    area.title = sheetnames[sIdx]
                    area.add_data(datas, titles_from_data=True)
                    area.set_categories(categories)
                    sheet.add_chart(area, "B" + str((maxr*4 + (15*3))))

                    area3d = AreaChart3D()
                    area3d.height = 10
                    area3d.legend = None
                    area3d.title = sheetnames[sIdx]
                    area3d.add_data(datas, titles_from_data=True)
                    area3d.set_categories(categories)
                    sheet.add_chart(area3d, "K" + str((maxr*4 + (15*3))))

                    break

                cIdx = 0
                if rIdx > 0:
                    maxc = len(row)
                    for cell in row:
                        if cIdx <= 0:
                            cIdx = cIdx + 1
                            continue

                        sr = sum_row(row, cell.value)
                        if sr is not None:
                            cell.value = sr
                            sa.append(sr)
                        else:
                            cell.value = cIdx
                        cIdx = cIdx + 1
                rIdx = rIdx + 1

        user_src = tdir + '/report_00.xlsx'
        book.save(user_src)

        dest = tdir + '/viewer.handsontable.html'
        css = 'static/css/handsontable.full.min.css'
        js = 'static/js/handsontable.full.min.js'
        book = pe.get_book(file_name=user_src, skip_hidden_sheets=False)
        # print(book)
        book.save_as(dest, readOnly=False, js_url=js, css_url=css)
    except IOError as e:
        print("IOError:" + str(e))
    except FileNotFoundError as e:
        print("FileNotFoundError:" + str(e))

    html = render_template('excel/viewer.html', excel_viewer=html)
    # delete_dir(tdir)
    return html
